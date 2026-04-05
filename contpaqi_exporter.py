# =====================================================================
# contpaqi_export_pro.py (ULTRA PRO v15 - Filtros y Totales Ajustados)
# Excel SolutionsV — Exportador de Nominas ContpaqI
# =====================================================================

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import subprocess
import re
import datetime
import pyodbc
import pandas as pd
import json
import os
import warnings
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# --- SILENCIADOR DE ADVERTENCIAS ---
# Pandas lanza alertas cuando no usamos SQLAlchemy.
# Como usamos pyodbc de manera directa por eficiencia, silenciamos estos mensajes.
warnings.filterwarnings('ignore', message='.*SQLAlchemy.*')
warnings.filterwarnings('ignore', message='.*DBAPI2.*')

CONFIG_FILE = "config.json"

# ──────────────────────────────────────────────
# ODBC DRIVERS: Prioridad de conexión
# ──────────────────────────────────────────────
ODBC_DRIVERS_PRIORITY = [
    "ODBC Driver 17 for SQL Server",
    "ODBC Driver 18 for SQL Server",
    "ODBC Driver 13 for SQL Server",
    "SQL Server Native Client 11.0",
    "SQL Server Native Client 10.0",
    "SQL Server",
]

def get_installed_sql_drivers():
    """Busca en el sistema Windows qué drivers de SQL Server están instalados."""
    return [d for d in pyodbc.drivers() if "sql server" in d.lower()]

# ──────────────────────────────────────────────
# COLORES Y ESTILOS DE LA INTERFAZ
# ──────────────────────────────────────────────
C = {
    "bg":        "#0F1117", 
    "surface":   "#1A1D27", 
    "card":      "#22263A",
    "border":    "#2E3250", 
    "accent":    "#3182DF", 
    "accent2":   "#5BA3FF",
    "success":   "#21B868", 
    "danger":    "#E74C3C", 
    "warning":   "#F4A62A",
    "text":      "#E8ECF0", 
    "text_dim":  "#7B8499", 
    "text_mute": "#454B63",
    "white":     "#FFFFFF", 
    "step_done": "#1A3A22", 
    "step_act":  "#0D2340",
    "step_idle": "#181B2A", 
    "log_bg":    "#0A0D14",
}
STEPS = ["1  Servidor", "2  Empresa", "3  Departamentos", "4  Exportar"]

# ──────────────────────────────────────────────
# CONFIGURACION Y DICCIONARIO DE CONTRASEÑAS 
# ──────────────────────────────────────────────
def load_config():
    """Carga el archivo config.json si existe."""
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r") as f: 
            return json.load(f)
    return {}

def save_config(data):
    """Guarda las configuraciones en config.json."""
    with open(CONFIG_FILE, "w") as f: 
        json.dump(data, f, indent=4)

# Diccionario amplio con contraseñas comunes (ContpaqI, Fiscalia, Reddit)
DEFAULT_PASSWORDS = [
    "Compac1", "compac", "Contpaqi1", "contpaqi", 
    "sa", "adminsa", "123456", "admin", "Admin123", 
    "master", "(en blanco)", "1234", "sql", "server", 
    "root", "COMPAC1", "CONTPAQI1", "temporal", "Temporal1"
]

def load_known_passwords():
    """Combina las contraseñas previas exitosas con el diccionario base."""
    cfg = load_config()
    saved = cfg.get("known_passwords", [])
    merged = list(saved)
    for p in DEFAULT_PASSWORDS:
        if p not in merged: 
            merged.append(p)
    return merged

def save_known_password(password):
    """Guarda una contraseña exitosa para que aparezca primero la próxima vez."""
    if not password: 
        return
    cfg = load_config()
    known = cfg.get("known_passwords", [])
    if password not in known: 
        known.insert(0, password)
    cfg["known_passwords"] = known
    save_config(cfg)

# ──────────────────────────────────────────────
# DETECCION DE RUTAS Y SDK DE CONTPAQI
# ──────────────────────────────────────────────
SDK_INSTALL_PATHS = [
    r"C:\ContpaqI\Nominas",
    r"C:\Program Files\Compac\Nominas",
    r"C:\Program Files (x86)\Compac\Nominas",
    r"C:\Compac\Nominas",
]

def detect_contpaqi_path():
    """
    Busca en el disco duro si existe alguna de las carpetas 
    típicas de instalación de ContpaqI Nóminas.
    """
    for path in SDK_INSTALL_PATHS:
        if os.path.isdir(path): 
            return path
    return None

def sdk_test_connection(install_path=None):
    """
    Retorna un mensaje advirtiendo que el SDK no es compatible 
    con arquitecturas de 64 bits de Python.
    """
    return False, None, (
        "NOTA IMPORTANTE: El SDK de ContpaqI es de 32-bits.\n"
        "Tu instalación de Python es de 64-bits, la conexión ha sido bloqueada.\n"
        "Por favor, usa la Autenticación SQL Server (es mucho más estable)."
    )

# ──────────────────────────────────────────────
# DETECCION AUTOMATICA DE SERVIDORES EN LA RED
# ──────────────────────────────────────────────
def _net_use_hosts():
    """Busca recursos compartidos en red que puedan ser el servidor."""
    hosts = set()
    try:
        r = subprocess.run(["net", "use"], capture_output=True, text=True, timeout=6, encoding="cp850", errors="replace")
        for m in re.finditer(r"\\\\([^\\]+)\\", r.stdout):
            if m.group(1).strip(): 
                hosts.add(m.group(1).strip())
    except: 
        pass
    return hosts

def detect_sql_servers():
    """Compila una lista de posibles servidores SQL (Prioridad a \\COMPAC)."""
    found = set()
    for host in _net_use_hosts():
        found.update([f"{host}\\COMPAC", host, f"{host}\\SQLEXPRESS", f"{host}\\MSSQLSERVER"])
    for d in ["(local)\\COMPAC", "(local)", "localhost\\COMPAC", "localhost", r".\COMPAC", r".\SQLEXPRESS"]:
        found.add(d)

    def key(s):
        sl = s.lower()
        if "compac" in sl and not sl.startswith("(") and not sl.startswith(".") and not sl.startswith("local"): 
            return (0, sl)
        if not sl.startswith("(") and not sl.startswith(".") and not sl.startswith("local"): 
            return (1, sl)
        return (2, sl)
    
    return sorted(found, key=key)

# ──────────────────────────────────────────────
# CONEXIONES SQL
# ──────────────────────────────────────────────
def _build_cs(server, driver, user=None, password=None, database=None):
    """Construye la cadena de conexión."""
    if user and password: 
        auth = f"UID={user};PWD={password};"
    else: 
        auth = "Trusted_Connection=yes;"
        
    cs = f"DRIVER={{{driver}}};SERVER={server};{auth}"
    
    if database: 
        cs += f"DATABASE={database};"
        
    return cs

def test_connection(server, user=None, password=None, timeout=8):
    """Prueba conectar al SQL Server iterando por todos los drivers disponibles."""
    installed = get_installed_sql_drivers()
    if not installed: 
        return False, None, "No hay ningún driver ODBC de SQL Server instalado."
        
    last_err = ""
    to_try = [d for d in ODBC_DRIVERS_PRIORITY if d in installed] + [d for d in installed if d not in ODBC_DRIVERS_PRIORITY]
    
    for driver in to_try:
        try:
            cs = _build_cs(server, driver, user, password)
            conn = pyodbc.connect(cs, timeout=timeout)
            conn.close()
            return True, driver, None
        except Exception as e:
            last_err = str(e)
            
    return False, None, last_err

def get_databases(server, driver, user=None, password=None):
    """Extrae las bases de datos excluyendo las del sistema."""
    cs = _build_cs(server, driver, user, password)
    conn = pyodbc.connect(cs, timeout=15)
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM sys.databases WHERE name NOT IN ('master', 'tempdb', 'model', 'msdb', 'ReportServer', 'ReportServerTempDB') ORDER BY name")
    rows = [r[0] for r in cursor.fetchall()]
    conn.close()
    return rows

def get_departments(server, database, driver, user=None, password=None):
    """Detecta si se usa NomDepartamento o nom10003 y extrae la lista."""
    cs = _build_cs(server, driver, user, password, database)
    conn = pyodbc.connect(cs, timeout=15)
    cursor = conn.cursor()
    
    cursor.execute("SELECT name FROM sys.tables WHERE name IN ('nom10003', 'NomDepartamento')")
    tables = [r[0].lower() for r in cursor.fetchall()]
    dept_table = 'nom10003' if 'nom10003' in tables else 'NomDepartamento'
    
    if not tables:
        conn.close()
        return pd.DataFrame(columns=['Id', 'Descripcion'])

    cursor.execute(f"SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = '{dept_table}'")
    cols = {r[0].lower(): r[0] for r in cursor.fetchall()}
    
    dept_id = cols.get('iddepartamento', cols.get('ciddepartamento'))
    dept_desc = cols.get('descripcion', cols.get('cnombredepartamento'))
    
    try:
        with warnings.catch_warnings():
            warnings.simplefilter('ignore', UserWarning)
            df = pd.read_sql(f"SELECT {dept_id} AS Id, {dept_desc} AS Descripcion FROM {dept_table} ORDER BY {dept_desc}", conn)
    except Exception:
        df = pd.DataFrame(columns=['Id', 'Descripcion'])
        
    conn.close()
    return df

# ──────────────────────────────────────────────
# MOTOR DE EXPORTACION Y LECTURA INTELIGENTE
# ──────────────────────────────────────────────
def safe_str(series):
    """Limpia los datos, previniendo que Pandas agregue decimales (.0) a los textos."""
    return series.fillna('').astype(str).str.replace(r'\.0$', '', regex=True).str.strip()

def export_data(server, database, output_path, only_active, selected_departments, driver, user=None, password=None):
    """Ejecuta la extracción de empleados detectando dinámicamente el esquema (Viejas vs Nuevas versiones)"""
    cs = _build_cs(server, driver, user, password, database)
    conn = pyodbc.connect(cs, timeout=30)
    cursor = conn.cursor()

    # 1. Identificar tablas
    cursor.execute("SELECT name FROM sys.tables WHERE name IN ('nom10001', 'NomEmpleado', 'nom10003', 'NomDepartamento')")
    existing_tables = [r[0].lower() for r in cursor.fetchall()]
    
    emp_table = 'nom10001' if 'nom10001' in existing_tables else 'NomEmpleado'
    dept_table = 'nom10003' if 'nom10003' in existing_tables else 'NomDepartamento'

    # 2. Extraer Diccionario de Columnas
    cursor.execute(f"SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = '{emp_table}'")
    cols_emp = {r[0].lower(): r[0] for r in cursor.fetchall()}

    cursor.execute(f"SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = '{dept_table}'")
    cols_dept = {r[0].lower(): r[0] for r in cursor.fetchall()}

    def get_real_col(diccionario, posibles):
        """Busca el nombre real de la columna en la base de datos actual."""
        for p in posibles:
            if p.lower() in diccionario: 
                return diccionario[p.lower()]
        return None

    # 3. Asignación Dinámica de Columnas
    emp_dept_id = get_real_col(cols_emp, ['idDepartamento', 'cIdDepartamento'])
    cat_dept_id = get_real_col(cols_dept, ['idDepartamento', 'cIdDepartamento'])
    dept_desc   = get_real_col(cols_dept, ['Descripcion', 'cNombreDepartamento'])
    
    codigo_emp  = get_real_col(cols_emp, ['CodigoEmpleado', 'cCodigoEmpleado'])
    diario_emp  = get_real_col(cols_emp, ['SueldoDiario', 'cSueldoDiario'])
    sdi_emp     = get_real_col(cols_emp, ['SalarioDiarioIntegrado', 'SueldoIntegrado', 'SBCPartFija', 'SBC', 'cSalarioDiarioIntegrado'])
    estatus_emp = get_real_col(cols_emp, ['EstadoEmpleado', 'cEstatus'])
    
    nombre_largo = get_real_col(cols_emp, ['NombreLargo'])
    nombre_corto = get_real_col(cols_emp, ['cNombre'])
    paterno      = get_real_col(cols_emp, ['cApellidoPaterno'])
    materno      = get_real_col(cols_emp, ['cApellidoMaterno'])

    # RFC y CURP Parts (El problema de ContpaqI)
    fecha_nac = get_real_col(cols_emp, ['FechaNacimiento', 'cFechaNacimiento'])
    fecha_rfc = get_real_col(cols_emp, ['FechaRFC', 'cFechaRFC'])

    if fecha_rfc:
        mid_rfc = f"ISNULL(E.{fecha_rfc}, '')"
    elif fecha_nac:
        # Convierte la Fecha de Nacimiento en formato AAMMDD (Ej: 920929)
        mid_rfc = f"CASE WHEN E.{fecha_nac} <= '1900-01-01' THEN '' ELSE ISNULL(CONVERT(varchar(6), E.{fecha_nac}, 12), '') END"
    else:
        mid_rfc = "''"

    rfc_p1 = get_real_col(cols_emp, ['RFC', 'cRFC'])
    rfc_p3 = get_real_col(cols_emp, ['Homoclave', 'cHomoclave'])
    
    curp_full = get_real_col(cols_emp, ['CURP', 'cCURP'])
    curpi     = get_real_col(cols_emp, ['CURPI', 'cCURPI'])
    curpf     = get_real_col(cols_emp, ['CURPF', 'cCURPF'])

    # SQL para armar RFC
    if rfc_p1 and rfc_p3:
        rfc_sql = f"LTRIM(RTRIM(ISNULL(E.{rfc_p1}, '') + {mid_rfc} + ISNULL(E.{rfc_p3}, '')))"
    elif rfc_p1:
        rfc_sql = f"LTRIM(RTRIM(E.{rfc_p1}))"
    else:
        rfc_sql = "''"

    # SQL para armar CURP
    if curpi and curpf:
        curp_sql = f"LTRIM(RTRIM(ISNULL(E.{curpi}, '') + {mid_rfc} + ISNULL(E.{curpf}, '')))"
    elif curp_full: 
        curp_sql = f"LTRIM(RTRIM(E.{curp_full}))"
    else:
        curp_sql = "''"

    # SQL para armar Nombre
    if nombre_largo:
        nombre_sql = f"E.{nombre_largo}"
    else:
        nombre_sql = f"LTRIM(RTRIM(ISNULL(E.{nombre_corto}, '') + ' ' + ISNULL(E.{paterno}, '') + ' ' + ISNULL(E.{materno}, '')))"

    # 4. Filtros
    where = []
    if only_active and estatus_emp:
        where.append(f"E.{estatus_emp} IN ('A', 'R')" if 'estado' in estatus_emp.lower() else f"E.{estatus_emp} = 1")
    if selected_departments and emp_dept_id:
        ids = ",".join(map(str, selected_departments))
        where.append(f"E.{emp_dept_id} IN ({ids})")
        
    wc = ("WHERE " + " AND ".join(where)) if where else ""

    # 5. La Consulta
    query = f"""
    SELECT
        E.{codigo_emp}           AS Codigo,
        {rfc_sql}                AS RFC,
        {curp_sql}               AS CURP,
        {nombre_sql}             AS Nombre,
        E.{diario_emp}           AS [Salario Diario],
        E.{sdi_emp}              AS SDI,
        D.{dept_desc}            AS Departamento
    FROM {emp_table} E
    LEFT JOIN {dept_table} D ON E.{emp_dept_id} = D.{cat_dept_id}
    {wc}
    ORDER BY E.{codigo_emp}
    """
    
    with warnings.catch_warnings():
        warnings.simplefilter('ignore', UserWarning)
        df = pd.read_sql(query, conn)
        
    conn.close()

    if df.empty:
        raise ValueError("La consulta no arrojó resultados. Verifica los filtros elegidos.")

    # 6. Creación del Excel y Estilos
    df.to_excel(output_path, index=False, engine="openpyxl")
    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "Empleados"

    BLUE = "3182DF"
    ROW_ALT = "EEF4FB"
    BC = "BDD0E8"
    thin = Border(**{s: Side(style="thin", color=BC) for s in ("left", "right", "top", "bottom")})

    # Cabeceras
    for cell in ws[1]:
        cell.fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Filas Cebras
    mr = ws.max_row
    for r in range(2, mr + 1):
        fc = ROW_ALT if r % 2 == 0 else "FFFFFF"
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color=fc, end_color=fc, fill_type="solid")
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center")

    # Ancho Automático de Columnas
    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 50)

    # --- NUEVO: Filtros Automáticos (AutoFilter) ---
    # Aplica el filtro desde A1 hasta la última columna/fila con datos.
    ultima_letra = ws.cell(row=1, column=ws.max_column).column_letter
    ws.auto_filter.ref = f"A1:{ultima_letra}{mr}"

    # --- CORRECCIÓN: Fila de Totales en Columna B y C (Columna A libre) ---
    tr = mr + 2
    
    # Se usa start=2 para que el texto "Total Empleados" inicie en B (Columna 2)
    for ci, (val, align, color) in enumerate([("Total Empleados:", "right", "0F1117"), (mr - 1, "center", BLUE)], start=2):
        cell = ws.cell(row=tr, column=ci, value=val)
        cell.font = Font(bold=True, color=color, name="Calibri", size=11)
        cell.fill = PatternFill(start_color="D6E4F5", end_color="D6E4F5", fill_type="solid")
        cell.alignment = Alignment(horizontal=align)
        cell.border = thin

    ws.freeze_panes = "A2"
    wb.save(output_path)
    return len(df)

# ──────────────────────────────────────────────
# INTERFAZ GRÁFICA PRINCIPAL
# ──────────────────────────────────────────────
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel SolutionsV \u00b7 Exportador Nominas ContpaqI")
        self.root.configure(bg=C["bg"])
        self.root.resizable(True, True)
        self.root.minsize(800, 820)

        self.config         = load_config()
        self._dept_ids      = []
        self._filter_active = True
        self._active_driver = None
        self._auth_mode     = "sql"      

        self._apply_styles()
        self._build_ui()
        self._center_window(860, 980)

        installed = get_installed_sql_drivers()
        if installed:
            self._log(f"Drivers ODBC encontrados: {', '.join(installed)}", "info")
        else:
            self._log("ERROR CRITICO: No hay ningún driver ODBC de SQL Server instalado.", "error")

        # Inicia la búsqueda automática de servidores en la red
        self.root.after(400, self._start_scan)

    def _apply_styles(self):
        """Define los estilos globales usando ttk.Style"""
        s = ttk.Style()
        s.theme_use("clam")
        
        s.configure("TFrame", background=C["bg"])
        s.configure("TLabelframe", background=C["card"], foreground=C["text_dim"], bordercolor=C["border"], relief="flat", padding=10)
        s.configure("TLabelframe.Label", background=C["card"], foreground=C["accent2"], font=("Segoe UI", 9, "bold"))
        
        s.configure("TCombobox", fieldbackground=C["surface"], background=C["surface"], foreground=C["text"], selectbackground=C["accent"], selectforeground=C["white"], bordercolor=C["border"], arrowcolor=C["accent2"], insertcolor=C["white"])
        s.map("TCombobox", fieldbackground=[("readonly", C["surface"])], foreground=[("readonly", C["text"])], insertcolor=[("focus", C["white"])])
        
        s.configure("TEntry", fieldbackground=C["surface"], foreground=C["text"], bordercolor=C["border"], insertcolor=C["text"])
        s.configure("Vertical.TScrollbar", background=C["surface"], troughcolor=C["bg"], arrowcolor=C["text_dim"], bordercolor=C["border"])
        s.configure("TProgressbar", troughcolor=C["surface"], background=C["success"], bordercolor=C["border"])

    def _build_ui(self):
        """Construye todos los elementos visuales de la aplicación paso a paso."""
        root = self.root

        # --- CABECERA SUPERIOR ---
        hdr = tk.Frame(root, bg=C["accent"], height=52)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        
        tk.Label(hdr, text="  \u2b21  Excel SolutionsV", bg=C["accent"], fg=C["white"], font=("Segoe UI", 14, "bold")).pack(side="left", padx=18, pady=10)
        tk.Label(hdr, text="Exportador de Nóminas ContpaqI (PRO v15)", bg=C["accent"], fg="#BFD9FF", font=("Segoe UI", 10)).pack(side="left", pady=10)

        # --- BARRA DE PASOS ---
        self._step_labels = []
        step_bar = tk.Frame(root, bg=C["surface"], height=46)
        step_bar.pack(fill="x")
        step_bar.pack_propagate(False)
        
        inner = tk.Frame(step_bar, bg=C["surface"])
        inner.pack(fill="both", expand=True)
        
        for i, label in enumerate(STEPS):
            if i > 0: 
                tk.Label(inner, text=" \u203a ", bg=C["surface"], fg=C["text_mute"], font=("Segoe UI", 14, "bold")).pack(side="left")
            lbl = tk.Label(inner, text=f"  {label}  ", bg=C["step_idle"], fg=C["text_mute"], font=("Segoe UI", 9, "bold"), padx=4, pady=5)
            lbl.pack(side="left", fill="y", padx=2, pady=5)
            self._step_labels.append(lbl)
            
        self._set_step(0)

        # --- CONTENEDOR PRINCIPAL CON SCROLL ---
        wrap = tk.Frame(root, bg=C["bg"])
        wrap.pack(fill="both", expand=True)
        
        canvas = tk.Canvas(wrap, bg=C["bg"], highlightthickness=0)
        vscroll = ttk.Scrollbar(wrap, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vscroll.set)
        vscroll.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        
        body = tk.Frame(canvas, bg=C["bg"], padx=16, pady=14)
        bid = canvas.create_window((0, 0), window=body, anchor="nw")
        
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(bid, width=e.width))
        body.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        # ==========================================
        # PASO 1: SERVIDOR
        # ==========================================
        s1 = ttk.LabelFrame(body, text="  \U0001f50c  Paso 1 \u2014 Servidor SQL")
        s1.pack(fill="x", pady=(0, 8))

        tk.Label(s1, text="Servidor:", bg=C["card"], fg=C["text_dim"], font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w", padx=(0, 8), pady=3)
        
        self.server_combo = ttk.Combobox(s1, width=34, state="normal", font=("Segoe UI", 10))
        self.server_combo.grid(row=0, column=1, sticky="ew", pady=3)
        self.server_combo.bind("<<ComboboxSelected>>", lambda e: self._on_server_pick())
        
        if self.config.get("server"): 
            self.server_combo.set(self.config["server"])

        self.scan_btn = self._mk_btn(s1, "\U0001f50d Buscar", self._start_scan, "secondary")
        self.scan_btn.grid(row=0, column=2, padx=(6, 0), pady=3)

        # Modos de Autenticación
        auth_frame = tk.Frame(s1, bg=C["card"])
        auth_frame.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(6, 2))
        
        tk.Label(auth_frame, text="Modo:", bg=C["card"], fg=C["text_dim"], font=("Segoe UI", 9)).pack(side="left", padx=(0, 8))

        self._auth_win_btn = tk.Button(auth_frame, text="🖥  Windows", command=lambda: self._set_auth_mode("windows"), bg=C["surface"], fg=C["text_mute"], activebackground=C["border"], activeforeground=C["white"], font=("Segoe UI", 9), relief="flat", cursor="hand2", padx=8, pady=3)
        self._auth_win_btn.pack(side="left", padx=(0, 3))

        self._auth_sql_btn = tk.Button(auth_frame, text="🔑  SQL Server", command=lambda: self._set_auth_mode("sql"), bg=C["accent"], fg=C["white"], activebackground="#255EAA", activeforeground=C["white"], font=("Segoe UI", 9, "bold"), relief="flat", cursor="hand2", padx=8, pady=3)
        self._auth_sql_btn.pack(side="left", padx=(0, 3))

        self._auth_sdk_btn = tk.Button(auth_frame, text="⚙  ContpaqI SDK", command=lambda: self._set_auth_mode("sdk"), bg=C["surface"], fg=C["text_mute"], activebackground=C["border"], activeforeground=C["white"], font=("Segoe UI", 9), relief="flat", cursor="hand2", padx=8, pady=3)
        self._auth_sdk_btn.pack(side="left")

        # Configuración SQL
        self._sql_auth_frame = tk.Frame(s1, bg=C["card"])
        self._sql_auth_frame.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(4, 2))

        tk.Label(self._sql_auth_frame, text="Usuario:", bg=C["card"], fg=C["text_dim"], font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w", padx=(0, 8), pady=3)
        
        self.user_entry = tk.Entry(self._sql_auth_frame, width=20, font=("Segoe UI", 10), bg=C["surface"], fg=C["text"], insertbackground=C["white"], relief="flat", highlightthickness=1, highlightcolor=C["accent"], highlightbackground=C["border"])
        self.user_entry.grid(row=0, column=1, sticky="w", pady=3)
        self.user_entry.insert(0, self.config.get("db_user", "sa"))

        tk.Label(self._sql_auth_frame, text="Password:", bg=C["card"], fg=C["text_dim"], font=("Segoe UI", 9)).grid(row=1, column=0, sticky="w", padx=(0, 8), pady=3)
        
        self.pass_entry = tk.Entry(self._sql_auth_frame, width=20, font=("Segoe UI", 10), bg=C["surface"], fg=C["text"], insertbackground=C["white"], show="\u2022", relief="flat", highlightthickness=1, highlightcolor=C["accent"], highlightbackground=C["border"])
        self.pass_entry.grid(row=1, column=1, sticky="w", pady=3)

        pass_ctrl = tk.Frame(self._sql_auth_frame, bg=C["card"])
        pass_ctrl.grid(row=1, column=2, sticky="w", padx=(8, 0))

        self._show_pass_var = tk.BooleanVar(value=False)
        tk.Checkbutton(pass_ctrl, text="Mostrar", variable=self._show_pass_var, command=self._toggle_pass_visibility, bg=C["card"], fg=C["text_dim"], activebackground=C["card"], selectcolor=C["surface"], font=("Segoe UI", 8), relief="flat").pack(side="left")

        self._pwd_menu_btn = tk.Menubutton(pass_ctrl, text="\u25be Probar comunes", bg=C["surface"], fg=C["accent2"], font=("Segoe UI", 8), relief="flat", cursor="hand2", activebackground=C["border"], activeforeground=C["white"])
        self._pwd_menu_btn.pack(side="left", padx=(6, 0))
        
        self._pwd_menu = tk.Menu(self._pwd_menu_btn, tearoff=0, bg=C["surface"], fg=C["text"], activebackground=C["accent"], activeforeground=C["white"], font=("Consolas", 9))
        self._pwd_menu_btn["menu"] = self._pwd_menu
        self._rebuild_password_menu()

        # El Boton Magico de Auto-Busqueda
        self._auto_pwd_btn = tk.Button(pass_ctrl, text="⚡ Auto-Buscar", command=self._auto_check_passwords, bg=C["accent"], fg=C["white"], activebackground="#255EAA", activeforeground=C["white"], font=("Segoe UI", 8, "bold"), relief="flat", cursor="hand2", padx=6, pady=2)
        self._auto_pwd_btn.pack(side="left", padx=(6, 0))

        # Configuración SDK (Visible solo en modo SDK)
        self._sdk_auth_frame = tk.Frame(s1, bg=C["card"])
        tk.Label(self._sdk_auth_frame, text="Ruta ContpaqI:", bg=C["card"], fg=C["text_dim"], font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w", padx=(0, 8), pady=3)
        
        self.sdk_path_entry = tk.Entry(self._sdk_auth_frame, width=32, font=("Segoe UI", 9), bg=C["surface"], fg=C["text"], insertbackground=C["white"], relief="flat", highlightthickness=1, highlightcolor=C["accent"], highlightbackground=C["border"])
        self.sdk_path_entry.grid(row=0, column=1, sticky="ew", pady=3)
        
        sdk_detected = detect_contpaqi_path()
        if sdk_detected: 
            self.sdk_path_entry.insert(0, sdk_detected)
            
        self._mk_btn(self._sdk_auth_frame, "📁", lambda: self.sdk_path_entry.insert(0, filedialog.askdirectory()), "ghost").grid(row=0, column=2, padx=(4, 0))

        # Indicador de conexión
        self.conn_indicator = tk.Label(s1, text="\u25cf  Sin probar", bg=C["card"], fg=C["text_mute"], font=("Segoe UI", 9, "bold"))
        self.conn_indicator.grid(row=3, column=1, sticky="w", pady=(4, 2))
        
        self._mk_btn(s1, "\u26a1 Probar conexion", self._test_conn, "ghost").grid(row=3, column=2, padx=(6, 0), pady=(4, 2))
        s1.columnconfigure(1, weight=1)

        # ==========================================
        # PASO 2: EMPRESA
        # ==========================================
        s2 = ttk.LabelFrame(body, text="  \U0001f3e2  Paso 2 \u2014 Empresa (base de datos)")
        s2.pack(fill="x", pady=(0, 8))
        
        tk.Label(s2, text="Empresa:", bg=C["card"], fg=C["text_dim"], font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w", padx=(0, 8))
        
        self.db_combo = ttk.Combobox(s2, width=34, state="readonly", font=("Segoe UI", 10))
        self.db_combo.grid(row=0, column=1, sticky="ew", pady=4)
        self.db_combo.bind("<<ComboboxSelected>>", lambda e: self._load_departments())
        
        self.load_db_btn = self._mk_btn(s2, "\u21bb Cargar empresas", self._load_databases, "primary")
        self.load_db_btn.grid(row=0, column=2, padx=(8, 0))
        s2.columnconfigure(1, weight=1)

        # ==========================================
        # PASO 3: DEPARTAMENTOS
        # ==========================================
        s3 = ttk.LabelFrame(body, text="  \U0001f4c2  Paso 3 \u2014 Departamentos (Clic = multiple)")
        s3.pack(fill="x", pady=(0, 8))
        
        lw = tk.Frame(s3, bg=C["card"])
        lw.pack(fill="x")
        
        self.dept_listbox = tk.Listbox(lw, selectmode=tk.MULTIPLE, bg=C["surface"], fg=C["text"], selectbackground=C["accent"], selectforeground=C["white"], font=("Consolas", 10), height=5, relief="flat", bd=0, activestyle="none", highlightthickness=1, highlightcolor=C["border"], highlightbackground=C["border"])
        self.dept_listbox.pack(side="left", fill="both", expand=True)
        
        sb = ttk.Scrollbar(lw, orient="vertical", command=self.dept_listbox.yview)
        sb.pack(side="right", fill="y")
        self.dept_listbox.configure(yscrollcommand=sb.set)

        # ==========================================
        # PASO 4: FILTROS
        # ==========================================
        s4 = ttk.LabelFrame(body, text="  \U0001f3af  Paso 4 \u2014 \u00bfQue empleados exportar?")
        s4.pack(fill="x", pady=(0, 8))
        
        tr2 = tk.Frame(s4, bg=C["card"])
        tr2.pack(fill="x")
        
        self._tog_activos = self._mk_toggle(tr2, "\u2714  Solo Activos", "Excluye bajas y suspendidos", True, lambda: self._set_filter(True))
        self._tog_activos.pack(side="left", fill="both", expand=True, padx=(0, 6), pady=4)
        
        self._tog_todos = self._mk_toggle(tr2, "\u229e  Todos los Empleados", "Incluye bajas, suspendidos e inactivos", False, lambda: self._set_filter(False))
        self._tog_todos.pack(side="left", fill="both", expand=True, pady=4)

        # ==========================================
        # PASO 5: ARCHIVO Y EXPORTACION
        # ==========================================
        s5 = ttk.LabelFrame(body, text="  \U0001f4be  Paso 5 \u2014 Archivo de salida")
        s5.pack(fill="x", pady=(0, 10))
        
        tk.Label(s5, text="Ruta:", bg=C["card"], fg=C["text_dim"], font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w", padx=(0, 8))
        
        self.file_entry = tk.Entry(s5, width=38, font=("Segoe UI", 10), bg=C["surface"], fg=C["text"], insertbackground=C["white"], relief="flat", highlightthickness=1, highlightcolor=C["accent"], highlightbackground=C["border"])
        self.file_entry.grid(row=0, column=1, sticky="ew", pady=4)
        
        self._mk_btn(s5, "\U0001f4c1 Examinar", self._select_file, "ghost").grid(row=0, column=2, padx=(8, 0))
        s5.columnconfigure(1, weight=1)

        self.progress = ttk.Progressbar(body, mode="indeterminate")

        self.gen_btn = tk.Button(body, text="  \u25b6   Generar Reporte a Excel", command=self._execute, bg=C["success"], fg=C["white"], activebackground="#189B52", activeforeground=C["white"], font=("Segoe UI", 12, "bold"), relief="flat", cursor="hand2", pady=14)
        self.gen_btn.pack(fill="x", pady=(0, 10))

        # ==========================================
        # PANEL DE LOGS (Con botones funcionales)
        # ==========================================
        log_hdr = tk.Frame(body, bg=C["bg"])
        log_hdr.pack(fill="x", pady=(0, 4))
        
        tk.Label(log_hdr, text="\U0001f4cb  Log de actividad", bg=C["bg"], fg=C["text_dim"], font=("Segoe UI", 9, "bold")).pack(side="left")
        
        self._mk_btn(log_hdr, "\U0001f4cb Copiar log", self._copy_log, "ghost").pack(side="right")
        self._mk_btn(log_hdr, "\u2715 Limpiar", self._clear_log, "ghost").pack(side="right", padx=(0, 6))

        self.log_box = tk.Text(body, bg=C["log_bg"], fg="#A8F0C6", font=("Consolas", 9), height=12, relief="flat", bd=0, state="disabled", wrap="word", highlightthickness=1, highlightcolor=C["border"], highlightbackground=C["border"])
        self.log_box.pack(fill="x", pady=(0, 6))
        
        self.log_box.tag_config("info", foreground="#5BA3FF")
        self.log_box.tag_config("success", foreground=C["success"])
        self.log_box.tag_config("warning", foreground=C["warning"])
        self.log_box.tag_config("error", foreground="#FF6B6B")
        self.log_box.tag_config("default", foreground="#A8F0C6")
        self.log_box.tag_config("ts", foreground=C["text_mute"])

        self._status_var = tk.StringVar(value="Listo.")
        
        # Asigna el nombre inicial al arrancar
        self._update_default_filename()

    # ──────────────────────────────────────────────
    # METODOS DE LOG Y ASISTENCIA UI
    # ──────────────────────────────────────────────
    def _copy_log(self):
        """Copia el contenido del recuadro de texto al portapapeles."""
        self.root.clipboard_clear()
        self.root.clipboard_append(self.log_box.get("1.0", "end"))

    def _clear_log(self):
        """Limpia todo el texto del recuadro de actividad."""
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    def _update_default_filename(self):
        """Asigna un nombre de archivo dinámico basándose en la fecha actual y el filtro."""
        hoy = datetime.datetime.now().strftime("%d-%m-%Y")
        estado = "Trabajadores Activos" if self._filter_active else "Todos los Trabajadores"
        sug_name = f"{estado} al {hoy}.xlsx"
        
        current_path = self.file_entry.get().strip()
        
        # Sobrescribe sólo si está vacío o si ya es un nombre sugerido previamente
        if not current_path or "Trabajadores" in current_path:
            dir_name = os.path.dirname(current_path) if current_path else ""
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, os.path.join(dir_name, sug_name) if dir_name else sug_name)

    def _select_file(self):
        """Abre el diálogo de guardar archivo, inyectando el nombre dinámico."""
        current_path = self.file_entry.get().strip()
        sug_name = os.path.basename(current_path) if current_path else "Reporte_Empleados.xlsx"
        
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel", "*.xlsx")],
            initialfile=sug_name, 
            title="Guardar reporte como..."
        )
        if path: 
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, path)

    def _set_auth_mode(self, mode: str):
        """Cambia entre la vista de Configuración Windows, SQL o SDK."""
        self._auth_mode = mode
        self._active_driver = None
        self.conn_indicator.configure(text="\u25cf  Sin probar", fg=C["text_mute"])
        
        dim  = {"bg": C["surface"], "fg": C["text_mute"], "font": ("Segoe UI", 9)}
        act  = {"bg": C["accent"],  "fg": C["white"],     "font": ("Segoe UI", 9, "bold")}
        
        self._auth_win_btn.configure(**(act if mode == "windows" else dim))
        self._auth_sql_btn.configure(**(act if mode == "sql" else dim))
        self._auth_sdk_btn.configure(**(act if mode == "sdk" else dim))
        
        if mode == "sql": 
            self._sql_auth_frame.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(4, 2))
            self._sdk_auth_frame.grid_remove()
        elif mode == "sdk": 
            self._sdk_auth_frame.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(4, 2))
            self._sql_auth_frame.grid_remove()
        else: 
            self._sql_auth_frame.grid_remove()
            self._sdk_auth_frame.grid_remove()
            
        self._log(f"Modo de autenticacion: {mode.upper()}", "info")

    def _rebuild_password_menu(self):
        """Reconstruye el menú desplegable de contraseñas guardadas."""
        self._pwd_menu.delete(0, tk.END)
        for pw in load_known_passwords():
            self._pwd_menu.add_command(
                label="(en blanco)" if pw == "" else pw, 
                command=lambda p=pw: (self.pass_entry.delete(0, tk.END), self.pass_entry.insert(0, p))
            )

    def _toggle_pass_visibility(self): 
        """Muestra u oculta los asteriscos de la contraseña."""
        self.pass_entry.configure(show="" if self._show_pass_var.get() else "\u2022")

    def _get_credentials(self): 
        """Retorna el usuario y contraseña del formulario SQL."""
        if self._auth_mode == "sql": 
            return self.user_entry.get().strip(), self.pass_entry.get()
        return None, None

    def _mk_btn(self, parent, text, command, kind="primary"):
        """Fábrica para crear botones consistentes."""
        cfg = {
            "primary": (C["accent"], C["white"], "#255EAA"), 
            "secondary": (C["surface"], C["accent2"], C["border"]), 
            "ghost": (C["card"], C["text_dim"], C["border"])
        }
        return tk.Button(parent, text=text, command=command, bg=cfg[kind][0], fg=cfg[kind][1], 
                         activebackground=cfg[kind][2], activeforeground=C["white"], 
                         font=("Segoe UI", 9, "bold" if kind == "primary" else "normal"), 
                         relief="flat", cursor="hand2", padx=10, pady=5)

    def _mk_toggle(self, parent, label, sublabel, selected, command):
        """Fábrica para crear las tarjetas clicables de los filtros."""
        frame = tk.Frame(parent, bg=C["card"], cursor="hand2", highlightthickness=2, highlightbackground=C["accent"] if selected else C["border"])
        t = tk.Label(frame, text=label, bg=C["card"], fg=C["accent2"] if selected else C["text_dim"], font=("Segoe UI", 11, "bold"), anchor="w", padx=10)
        t.pack(fill="x", pady=(8, 1))
        s = tk.Label(frame, text=sublabel, bg=C["card"], fg=C["text_dim"] if selected else C["text_mute"], font=("Segoe UI", 8), anchor="w", padx=12)
        s.pack(fill="x", pady=(0, 8))
        frame._title_lbl = t
        frame._sub_lbl = s
        for w in (frame, t, s): 
            w.bind("<Button-1>", lambda e: command())
        return frame

    def _set_filter(self, only_active: bool):
        """Cambia el estilo visual de las tarjetas de filtro cuando se les da clic."""
        self._filter_active = only_active
        for frame, is_act in [(self._tog_activos, True), (self._tog_todos, False)]:
            sel = (only_active == is_act)
            frame.configure(highlightbackground=C["accent"] if sel else C["border"])
            frame._title_lbl.configure(fg=C["accent2"] if sel else C["text_dim"])
            frame._sub_lbl.configure(fg=C["text_dim"] if sel else C["text_mute"])
            
        self._update_default_filename()

    def _set_step(self, step):
        """Actualiza el color de la barra de pasos en la parte superior."""
        for i, lbl in enumerate(self._step_labels):
            if i < step: 
                lbl.configure(bg=C["step_done"], fg=C["success"])
            elif i == step: 
                lbl.configure(bg=C["step_act"], fg=C["accent2"])
            else: 
                lbl.configure(bg=C["step_idle"], fg=C["text_mute"])

    def _center_window(self, w, h):
        """Centra la ventana principal en el monitor del usuario."""
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _log(self, msg, level="default"):
        """Imprime un mensaje formateado en el recuadro negro de Actividad."""
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        prefix = {"info": "[INFO]", "success": "[OK]  ", "warning": "[WARN]", "error": "[ERR] "}.get(level, "[LOG] ")
        
        self.log_box.configure(state="normal")
        self.log_box.insert("end", f"{ts}  ", "ts")
        self.log_box.insert("end", f"{prefix}  ", level)
        self.log_box.insert("end", msg + "\n", level)
        self.log_box.configure(state="disabled")
        self.log_box.see("end")

    # ──────────────────────────────────────────────
    # RUTINAS PRINCIPALES (Threads)
    # ──────────────────────────────────────────────
    def _auto_check_passwords(self):
        """Fuerza bruta amigable: Itera las contraseñas comunes contra SQL Server."""
        server = self.server_combo.get().strip()
        user = self.user_entry.get().strip()
        
        if not server or not user: 
            messagebox.showwarning("Aviso", "Selecciona el servidor y escribe el usuario SQL primero.")
            return
            
        self._log("Iniciando auto-busqueda de contraseñas...", "info")
        self.conn_indicator.configure(text="● Buscando password...", fg=C["warning"])
        self._auto_pwd_btn.configure(state="disabled")

        def _work():
            for pwd in load_known_passwords():
                pwd_actual = "" if pwd == "(en blanco)" else pwd
                self.root.after(0, self._log, f"Probando: '{pwd}'...", "default")
                
                ok, driver, err = test_connection(server, user, pwd_actual, timeout=3)
                if ok: 
                    self.root.after(0, self._auto_check_success, pwd_actual, driver)
                    return
                    
            self.root.after(0, self._auto_check_fail)
            
        threading.Thread(target=_work, daemon=True).start()

    def _auto_check_success(self, pwd, driver):
        """Se ejecuta cuando el autocompletar encuentra la contraseña correcta."""
        self._active_driver = driver
        self.pass_entry.delete(0, tk.END)
        self.pass_entry.insert(0, pwd)
        save_known_password(pwd)
        self._rebuild_password_menu()
        
        self.conn_indicator.configure(text=f"● Conectado ✔ [{driver}]", fg=C["success"])
        self._set_step(1)
        self._auto_pwd_btn.configure(state="normal")
        self._log(f"Contraseña correcta encontrada y guardada.", "success")

    def _auto_check_fail(self):
        """Se ejecuta si la fuerza bruta no logra dar con la clave."""
        self.conn_indicator.configure(text="● Fallo", fg=C["danger"])
        self._auto_pwd_btn.configure(state="normal")
        self._log("Ninguna contraseña del diccionario funcionó.", "error")

    def _start_scan(self):
        """Inicia el escaneo de red en segundo plano para no congelar la app."""
        self.scan_btn.configure(text="\u23f3 Buscando...", state="disabled")
        threading.Thread(target=lambda: self.root.after(0, self._scan_done, detect_sql_servers()), daemon=True).start()

    def _scan_done(self, servers):
        """Recibe la lista de servidores tras el escaneo."""
        self.server_combo["values"] = list(servers)
        if servers: 
            self.server_combo.set(list(servers)[0])
        self.scan_btn.configure(text="\U0001f50d Buscar", state="normal")

    def _on_server_pick(self):
        """Resetea el estado si el usuario cambia el servidor del dropdown."""
        self._set_step(0)
        self._active_driver = None
        self.conn_indicator.configure(text="\u25cf  Sin probar", fg=C["text_mute"])

    def _test_conn(self):
        """Envía la prueba de conexión a SQL o al SDK de ContpaqI."""
        server = self.server_combo.get().strip()
        
        if self._auth_mode == "sdk":
            threading.Thread(target=lambda: self.root.after(0, self._test_done_sdk, *sdk_test_connection(None)), daemon=True).start()
            return
            
        user, password = self._get_credentials()
        threading.Thread(target=lambda: self.root.after(0, self._test_done, server, *test_connection(server, user, password)), daemon=True).start()

    def _test_done(self, server, ok, driver, err):
        """Maneja el resultado de la conexión SQL."""
        if ok:
            self._active_driver = driver
            self.conn_indicator.configure(text=f"\u25cf  Conectado \u2714", fg=C["success"])
            self._set_step(1)
            
            if self._auth_mode == "sql": 
                save_known_password(self.pass_entry.get())
                self._rebuild_password_menu()
        else: 
            self.conn_indicator.configure(text="\u25cf  Fallo", fg=C["danger"])
            self._log(f"Error al conectar: {err}", "error")

    def _test_done_sdk(self, ok, info, err):
        """Maneja el resultado de la conexión SDK."""
        self.conn_indicator.configure(text="\u25cf  SDK fallo", fg=C["danger"])
        self._log(f"Error SDK: {err}", "error")

    def _load_databases(self):
        """Inicia la extracción de las bases de datos (Empresas)."""
        server = self.server_combo.get().strip()
        user, password = self._get_credentials()
        threading.Thread(target=lambda: self.root.after(0, self._dbs_loaded, get_databases(server, self._active_driver, user, password)), daemon=True).start()

    def _dbs_loaded(self, dbs):
        """Llena el dropdown de Empresas tras consultar SQL."""
        if dbs:
            self.db_combo["values"] = dbs
            self.db_combo.current(0)
            self._set_step(2)
            self._load_departments()
        else: 
            messagebox.showwarning("Aviso", "No se encontraron bases de datos válidas.")

    def _load_departments(self):
        """Inicia la extracción de los departamentos de la empresa seleccionada."""
        server = self.server_combo.get().strip()
        db = self.db_combo.get().strip()
        user, password = self._get_credentials()
        
        self.dept_listbox.delete(0, tk.END)
        self._dept_ids = []
        
        threading.Thread(target=lambda: self.root.after(0, self._depts_loaded, get_departments(server, db, self._active_driver, user, password)), daemon=True).start()

    def _depts_loaded(self, df):
        """Dibuja los departamentos en la lista interactiva."""
        for _, row in df.iterrows():
            self._dept_ids.append(row["Id"])
            self.dept_listbox.insert(tk.END, f"  {str(row['Id']):>4}  \u2502  {row['Descripcion']}")
        self._set_step(3)

    def _execute(self):
        """Dispara el hilo final que extrae los empleados y arma el Excel."""
        server = self.server_combo.get().strip()
        db = self.db_combo.get().strip()
        file_path = self.file_entry.get().strip()
        user, password = self._get_credentials()
        
        sel_idx = self.dept_listbox.curselection()
        sel_ids = [self._dept_ids[i] for i in sel_idx] if sel_idx else None
        
        self.progress.pack(fill="x", pady=(0, 6), before=self.gen_btn)
        self.progress.start(10)
        self.gen_btn.configure(state="disabled", text="  \u23f3  Generando...")

        def _work():
            try:
                n = export_data(server, db, file_path, self._filter_active, sel_ids, self._active_driver, user, password)
                self.root.after(0, self._export_done, n, file_path)
            except Exception as e:
                self.root.after(0, self._on_error, str(e))
                
        threading.Thread(target=_work, daemon=True).start()

    def _export_done(self, n_rows, file_path):
        """Termina el proceso de Excel limpiamente y muestra el cartel de éxito."""
        self.progress.stop()
        self.progress.pack_forget()
        self.gen_btn.configure(state="normal", text="  \u25b6   Generar Reporte a Excel")
        self._set_step(4)
        
        messagebox.showinfo("Listo \u2705", f"Exportacion exitosa\n\nEmpleados: {n_rows}\nArchivo: {file_path}")
        
        if messagebox.askyesno("\u00bfAbrir?", "\u00bfAbrir el Excel ahora?"): 
            os.startfile(file_path)

    def _on_error(self, msg):
        """Captura cualquier colapso de la consulta SQL y lo muestra al usuario."""
        self.progress.stop()
        self.progress.pack_forget()
        self.gen_btn.configure(state="normal", text="  \u25b6   Generar Reporte a Excel")
        self._log(f"ERROR: {msg}", "error")
        messagebox.showerror("Error", f"Ocurrio un error:\n\n{msg}")

if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()